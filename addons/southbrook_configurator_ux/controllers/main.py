# SPDX-License-Identifier: LGPL-3.0-only
"""Southbrook Configurator UX v2 — JSON-RPC endpoints.

Phase 2a:
    POST /southbrook/api/configurator/state
        Read-only. Returns the product's attribute_lines + values +
        price_extra + base price + a product.config.session id (created
        or reused).

Phase 2c:
    POST /southbrook/api/configurator/select
        Updates the session value_ids, returns server-resolved price +
        weight + disabled_value_ids from the OCA product.config.line
        rule engine. Replaces the hardcoded isValueDisabled() rules
        that shipped in Phase 2b.

    POST /southbrook/api/configurator/commit
        Materialises the variant via session.create_get_variant() +
        adds it to the user's draft sale.order, then moves the session
        to state='done'. Per the Phase-2 cart-target decision: A —
        Order Builder, not website_sale cart. Returns a redirect URL
        to /my/southbrook/order-builder/<id> so the client can
        navigate.

Phase 4 (this commit):
    GET  /southbrook/api/import/template
        Returns an xlsx file with the live attribute / category / UoM
        vocabulary baked in as dropdown sources + example rows. Replaces
        the client-side CSV download the bulk-tools button shipped with
        in Phases 1-3. Backend-only via auth='user' + internal-user check
        (portal users can't see the bulk-tools bar anyway).

    POST /southbrook/api/import/preview
        Accepts a multipart xlsx upload. Parses the PRODUCTS sheet,
        validates each row against the live vocab + the upsert
        invariants, returns a per-row {row, sheet, status, errors[]}
        list. Writes NOTHING — purely a dry-run preview. Backend-only.

    POST /southbrook/api/import/commit
        Same parsing + validation as /preview, but writes the VALID rows
        inside a single transaction. Requires `confirm: true` in the
        payload (refuses otherwise — explicit human gate per the Phase 4
        stop-point). Upserts product.template by default_code. Returns
        per-row commit log + summary counters. Backend-only.

    v1 of the import pipeline handles the PRODUCTS sheet only. The
    template's ATTRIBUTE_LINES / ATTRIBUTE_VALUES / BOM_* / HARDWARE_BOM /
    ACCESSORIES sheets are recognised but skipped with a "deferred to v2"
    info row in the preview output. The endpoint contract is shaped to
    accept multi-sheet payloads from day one so the v2 expansion lands
    without breaking clients.

Auth model: auth='public' + type='json'. The /shop/<slug> page is
publicly accessible (it's a catalog page), so anonymous visitors can
configure too. JSON-RPC routes don't require CSRF tokens. Sessions
for anonymous visitors live under base.public_user with the standard
OCA cleanup TTL.

No schema changes, no writes outside of product.config.session — which
the OCA module is the canonical custodian of. We sudo() the session
ops because public users don't have direct write ACL on that model,
but the scope is tight: search-or-create then attach to the request
user (not to a different user).
"""
import io
import json
import logging

from odoo import http
from odoo.exceptions import AccessError, UserError, ValidationError
from odoo.http import request

_logger = logging.getLogger(__name__)

# Bool coercion accepted by the importer for fields like sale_ok,
# is_published, manufacture_route, etc. Anything else is rejected as
# a row-level error.
_TRUE_TOKENS = {"true", "yes", "1", "y", "t"}
_FALSE_TOKENS = {"false", "no", "0", "n", "f", ""}

# Allowed values for the southbrook_category Selection field. Mirrors
# the field definition in southbrook_estimating.product_template.
_SOUTHBROOK_CATEGORIES = {"Wall", "Base", "Drawer", "Tall", "Vanity", "Extras"}

# Allowed values for southbrook_icon_key. Validated against the JS
# CABINET_ICONS map at configurator.esm.js — keep these in sync.
_SOUTHBROOK_ICON_KEYS = {
    "wall1", "wall2", "base1", "base2", "drawer", "sink",
    "pantry", "oven", "corner", "vanity", "extra", "worktop",
}

# Required PRODUCTS columns. The importer refuses the row if any is
# blank. uom_id has a default of "Units" applied before this check.
_PRODUCTS_REQUIRED = ("default_code", "name", "type", "internal_category")

# Recognised sheets in v1. PRODUCTS is processed; the rest are skipped
# with a "deferred to v2" status row in the preview so callers can see
# the importer DID notice them.
_V1_SHEET = "PRODUCTS"
_V2_SHEETS = ("ATTRIBUTE_LINES", "ATTRIBUTE_VALUES", "BOM_HEADERS",
              "BOM_LINES", "HARDWARE_BOM", "ACCESSORIES")
# Reference sheets that the template ships with; the importer reads
# nothing from these (they're dropdown sources for the spreadsheet
# itself), but it surfaces them as INFO rows in the preview so a
# misnamed sheet is easy to spot.
_REF_SHEETS = ("Instructions", "REF_CATEGORIES", "REF_ATTRIBUTES",
               "REF_UOM", "REF_FIELDS", "VERSION_STAMP")


# Logical attribute groups for the configurator's right pane. Matches
# the Phase-1 prototype's 4-group layout, keyed by canonical attribute
# name. Attributes on the template that don't appear here fall through
# into the "Other" group at the end of the list so a future new
# attribute doesn't silently disappear from the UI.
#
# Phase 3+ replaces this with a configurable grouping (likely a small
# ir.config_parameter table or a new field on product.attribute). For
# Phase 2 we keep the visual contract with the prototype.
ATTRIBUTE_GROUPS = [
    ("Size & Layout",         ["Width", "Door Count"]),
    ("Series & Materials",    ["Series", "Box Material", "Door Style"]),
    ("Finish & Construction", ["Finish", "Hinge Side", "Finished Sides", "Gables"]),
    ("Hardware & Add-ons",    ["Handle", "Accessories"]),
]


class SouthbrookConfiguratorAPI(http.Controller):
    """JSON-RPC endpoints for the v2 configurator UX."""

    # ------------------------------------------------------------------
    # /state — read-only product + attributes + session payload.
    # ------------------------------------------------------------------
    @http.route(
        "/southbrook/api/configurator/state",
        type="json",
        auth="public",
        methods=["POST"],
        website=True,
    )
    def configurator_state(self, product_tmpl_id=None, **kw):
        """Return everything the OWL configurator needs to render.

        Args:
            product_tmpl_id: int — the product.template id resolved by
                the OWL bootstrap from the page's mount-point div.

        Response (success):
            {
              ok: True,
              product: {tmpl_id, name, sku, list_price, currency:
                        {symbol, position, decimal_places, name}},
              session_id: int,
              base_price: float,
              groups: [{title: str, attribute_ids: [int, ...]}, ...],
              attributes: {
                "<attribute_id>": {
                  name: str, display_type: str, sequence: int,
                  required: bool,
                  values: [
                    {id, name, price_extra, html_color, sequence},
                    ...
                  ]
                },
                ...
              },
              selected_value_ids: [int, ...]   // existing picks if any
            }

        Response (error):
            {ok: False, error: "<code>", message: "<human-readable>"}
        """
        if not product_tmpl_id:
            return {"ok": False, "error": "missing_product_tmpl_id",
                    "message": "product_tmpl_id is required"}

        try:
            tmpl_id_int = int(product_tmpl_id)
        except (TypeError, ValueError):
            return {"ok": False, "error": "bad_product_tmpl_id",
                    "message": f"product_tmpl_id must be int, got {product_tmpl_id!r}"}

        # Sudo because the OCA configurator option-products attached to
        # attribute values aren't website_published; without sudo the
        # public ir.rule "Public product template" 403s on the
        # template's attribute_line_ids traversal. Same defect /
        # workaround as the OCA WebsiteSale.product override at
        # addons/website_product_configurator/controllers/main.py:135-156.
        # Scope discipline: only the read path is sudo'd; we never
        # write through this sudo recordset.
        Tmpl = request.env["product.template"].sudo()
        tmpl = Tmpl.browse(tmpl_id_int).exists()
        if not tmpl:
            return {"ok": False, "error": "product_not_found",
                    "message": f"No product.template with id={tmpl_id_int}"}
        if not tmpl.config_ok:
            return {"ok": False, "error": "not_configurable",
                    "message": f"{tmpl.display_name} is not configurable "
                               f"(config_ok=False)"}

        # Resolve or create a config session for this user + template.
        # Public visitors get a session attached to base.public_user; the
        # OCA module's standard cleanup cron sweeps stale public sessions.
        session = self._get_or_create_session(tmpl)

        # Build the attribute payload. Iterate attribute_line_ids in
        # sequence order so the UI renders attributes consistently.
        attributes = {}
        for line in tmpl.attribute_line_ids.sorted(key=lambda l: l.attribute_id.sequence):
            attr = line.attribute_id
            # Per-template attribute values carry the price_extra +
            # html_color overrides. The OCA model auto-creates one row
            # per (template, value) when an attribute_line is added.
            ptav_by_value_id = {
                ptav.product_attribute_value_id.id: ptav
                for ptav in line.product_template_value_ids
            }
            values = []
            for val in line.value_ids.sorted(key=lambda v: v.sequence):
                ptav = ptav_by_value_id.get(val.id)
                values.append({
                    "id": val.id,
                    "name": val.name,
                    "sequence": val.sequence,
                    "price_extra": float(ptav.price_extra) if ptav else 0.0,
                    "html_color": ptav.html_color if ptav and ptav.html_color
                                  else (val.html_color or None),
                    # Phase 2c will set this from product.config.line
                    # rule evaluation; for now every value starts enabled.
                    "disabled": False,
                })
            attributes[str(attr.id)] = {
                "name": attr.name,
                "display_type": attr.display_type,
                "sequence": attr.sequence,
                # OCA's required is on the attribute_line, not the
                # attribute, so we expose the line-level flag.
                "required": bool(line.required) if hasattr(line, "required") else False,
                "values": values,
            }

        # Map ATTRIBUTE_GROUPS canonical names to live attribute ids.
        # Attributes present on the template but not in any group fall
        # through into an "Other" group at the end so they're always
        # visible.
        group_payload = []
        all_attr_ids = {a.attribute_id.id: a.attribute_id.name
                        for a in tmpl.attribute_line_ids}
        used_ids = set()
        for title, names in ATTRIBUTE_GROUPS:
            ids = [aid for aid, nm in all_attr_ids.items() if nm in names]
            if ids:
                group_payload.append({"title": title, "attribute_ids": ids})
                used_ids.update(ids)
        leftover = [aid for aid in all_attr_ids if aid not in used_ids]
        if leftover:
            group_payload.append({"title": "Other", "attribute_ids": leftover})

        # Selected values: anything the session already carries — lets
        # a customer reload the page and pick up where they left off.
        selected = session.value_ids.ids if session else []

        # Resolve the SKU from the lowest-id variant (the v19
        # template.default_code-blanks-on-multi-variant gotcha).
        sku = ""
        for variant in tmpl.product_variant_ids.sorted("id"):
            if variant.default_code:
                sku = variant.default_code
                break
        if not sku and tmpl.default_code:
            sku = tmpl.default_code

        # Website currency for the UI's $ formatter.
        Website = request.env["website"].sudo()
        website = (Website.get_current_website()
                   if hasattr(Website, "get_current_website") else Website)
        currency = ((website and website.currency_id)
                    or request.env.company.currency_id)

        return {
            "ok": True,
            "product": {
                "tmpl_id": tmpl.id,
                "sku": sku,
                "name": tmpl.name,
                "list_price": float(tmpl.list_price),
                "currency": {
                    "symbol": currency.symbol or "$",
                    "position": currency.position or "before",
                    "decimal_places": currency.decimal_places or 2,
                    "name": currency.name or "CAD",
                },
            },
            "session_id": session.id if session else None,
            "base_price": float(tmpl.list_price),
            "groups": group_payload,
            "attributes": attributes,
            "selected_value_ids": selected,
        }

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_or_create_session(self, tmpl):
        """Search for an existing session for (user, template); create if absent.

        Sessions are tied to res.users. Public visitors land on
        base.public_user; portal / internal users get their own
        session. The OCA configurator's cleanup cron handles
        abandoned public sessions on a TTL.
        """
        Session = request.env["product.config.session"]
        # Search as the actual user first — if they have draft sessions
        # we want to honour their ACL. If the search returns empty AND
        # we have to create, that's where we sudo (so public users can
        # create their own session row).
        existing = Session.search([
            ("product_tmpl_id", "=", tmpl.id),
            ("user_id", "=", request.env.user.id),
            ("state", "=", "draft"),
        ], limit=1, order="create_date desc")
        if existing:
            return existing
        return Session.sudo().create({
            "product_tmpl_id": tmpl.id,
            "user_id": request.env.user.id,
        })

    # ------------------------------------------------------------------
    # /select — apply picks, return server-resolved price + disabled set.
    # ------------------------------------------------------------------
    @http.route(
        "/southbrook/api/configurator/select",
        type="json",
        auth="public",
        methods=["POST"],
        website=True,
    )
    def configurator_select(self, session_id=None, value_ids=None, **kw):
        """Update the session's value_ids and return server-resolved
        price + weight + disabled_value_ids (from the OCA rule engine).

        Client contract: send the COMPLETE current pick set as
        `value_ids` — the server resolves diffs internally. Sending a
        smaller set than last time means the missing attributes get
        cleared. Sending an unrelated value_id (not on the template)
        gets silently skipped.

        Args:
            session_id: int — the session created by /state
            value_ids: [int] — current picks, one value per attribute
                       at most

        Response (success):
            {
              ok: True,
              selected_value_ids: [int],     // what the server settled on
              price: float,                  // base + sum(price_extras)
              weight: float,                 // computed by OCA
              disabled_value_ids: [int],     // values forbidden by rules
              warnings: [str]                // reserved for soft messages
            }

        Response (error):
            {ok: False, error: "<code>", message: "<text>"}
        """
        if not session_id:
            return {"ok": False, "error": "missing_session_id"}
        if not isinstance(value_ids, list):
            return {"ok": False, "error": "value_ids_must_be_list",
                    "message": "value_ids must be a JSON array of ints"}

        try:
            session = self._authorize_session(int(session_id))
        except ValueError:
            return {"ok": False, "error": "bad_session_id"}
        if isinstance(session, dict):
            return session  # error dict bubbled up from the helper

        if session.state != "draft":
            return {"ok": False, "error": "session_locked",
                    "message": "This configuration was already committed."}

        tmpl = session.product_tmpl_id

        # Build {attribute_id: value_id} dict from the flat value_ids
        # list. Resolve each value's attribute via the global
        # product.attribute.value record. Unknown values are skipped;
        # if two values land for the same attribute (client bug), the
        # later one wins via dict semantics.
        Value = request.env["product.attribute.value"].sudo()
        attr_val_dict = {}
        for vid in value_ids:
            try:
                vid_int = int(vid)
            except (TypeError, ValueError):
                continue
            val = Value.browse(vid_int).exists()
            if val:
                attr_val_dict[val.attribute_id.id] = vid_int

        # Attributes on the template that aren't in the new picks get
        # explicitly cleared so OCA's update_config removes their old
        # value rather than carrying it over.
        for line in tmpl.attribute_line_ids:
            aid = line.attribute_id.id
            if aid not in attr_val_dict:
                attr_val_dict[aid] = []

        try:
            session.sudo().update_config(attr_val_dict)
        except (UserError, ValidationError) as exc:
            return {"ok": False, "error": "rule_blocked",
                    "message": getattr(exc, "args", [str(exc)])[0]}

        # Compute the disabled set: every value the template exposes
        # minus the available subset given current picks.
        all_val_ids = list({
            vid
            for line in tmpl.attribute_line_ids
            for vid in line.value_ids.ids
        })
        try:
            available_ids = session.sudo().values_available(
                check_val_ids=list(all_val_ids),
            )
            available_ids = [int(v) for v in available_ids]
        except Exception:                                  # noqa: BLE001
            # Belt-and-braces: if values_available barfs (rare; usually
            # only if a malformed rule landed in the DB), fall back to
            # "nothing disabled" rather than failing the whole pick.
            _logger.warning(
                "values_available raised on session %s; treating all "
                "values as enabled this round.",
                session.id, exc_info=True,
            )
            available_ids = list(all_val_ids)
        disabled_ids = set(all_val_ids) - set(available_ids)

        # Filter out "premature" disables. OCA's values_available
        # treats a value as unavailable whenever NO config.line domain
        # matches the current picks — including the empty-picks case.
        # At page-load with nothing chosen, every value with any rule
        # whose domain depends on Series (or any other attribute) gets
        # flagged disabled, because no Series has been picked yet to
        # satisfy the "Series in [...]" condition. The customer sees
        # ALL Box Material chips greyed and can't progress.
        #
        # The customer's mental model: a chip is disabled only if a
        # value I already PICKED makes it impossible. If the trigger
        # attribute hasn't been picked yet, the chip should still be
        # selectable. So for every disabled value, walk its restricting
        # rules: keep it disabled only if at least one rule has a
        # trigger attribute the user HAS picked AND the user's pick
        # doesn't satisfy that rule's allow list. Drop disability for
        # rules whose trigger attributes are still unset.
        picked_value_ids_by_attr = {
            aid: vid for aid, vid in attr_val_dict.items()
            if vid and vid != []
        }
        # Convert single-int picks to sets for the membership test
        # below. Treat list picks (which the controller uses to clear
        # an attribute via update_config) as "nothing picked".
        picked_set_by_attr = {
            aid: {vid if isinstance(vid, int) else None}
            for aid, vid in picked_value_ids_by_attr.items()
            if isinstance(vid, int)
        }
        refined_disabled = set()
        for vid in disabled_ids:
            # Find every config.line that mentions this value. These
            # are the rules that can FORBID v under the wrong picks.
            rules = tmpl.config_line_ids.filtered(
                lambda r, _vid=vid: _vid in r.value_ids.ids)
            for rule in rules:
                # A rule has one or more domain.line records (attribute,
                # condition, value_ids). For each:
                rule_blocks = False
                for dl in rule.domain_id.domain_line_ids:
                    trigger_attr_id = dl.attribute_id.id
                    if trigger_attr_id not in picked_set_by_attr:
                        # User hasn't picked the trigger attribute yet
                        # — this rule's domain isn't actively blocking.
                        continue
                    user_picks = picked_set_by_attr[trigger_attr_id]
                    allowed = set(dl.value_ids.ids)
                    if dl.condition == "in":
                        # Rule says trigger ∈ allowed; if user's pick
                        # isn't in allowed, the rule blocks.
                        if not (user_picks & allowed):
                            rule_blocks = True
                            break
                    else:  # condition == "not in"
                        # Rule says trigger ∉ allowed; if user's pick
                        # IS in allowed, the rule blocks.
                        if user_picks & allowed:
                            rule_blocks = True
                            break
                if rule_blocks:
                    refined_disabled.add(vid)
                    break
        disabled_ids = sorted(refined_disabled)

        return {
            "ok": True,
            "selected_value_ids": session.value_ids.ids,
            "price": float(session.price or 0.0),
            "weight": float(getattr(session, "weight", 0.0) or 0.0),
            "disabled_value_ids": disabled_ids,
            "warnings": [],
        }

    # ------------------------------------------------------------------
    # /commit — materialise variant + add to user's draft sale.order.
    # ------------------------------------------------------------------
    @http.route(
        "/southbrook/api/configurator/commit",
        type="json",
        auth="public",
        methods=["POST"],
        website=True,
    )
    def configurator_commit(self, session_id=None, order_id=None, **kw):
        """Materialise a product.product variant from the session +
        add it as a sale.order.line on the visitor's draft order.

        Public visitors get login_required (anonymous quotes aren't
        supported in the Southbrook flow per the existing customer-
        flow controller). They're sent to /web/signup with a return
        URL back to the current configurator page so they can finish
        what they started after signup.

        Per the Phase-2 cart-target decision: A — Order Builder.
        On success the client navigates to
        /my/southbrook/order-builder/<order_id> (the redirect field
        in the response).

        Args:
            session_id: int — the configured session
            order_id: int — optional. If supplied, the variant is
                      appended to that order (which must belong to
                      the same user). If absent, the importer reuses
                      the user's most-recent draft sale.order, or
                      creates a new draft if none exists.

        Response (success):
            {
              ok: True,
              variant_id: int,
              order_id: int,
              order_line_id: int,
              redirect: "/my/southbrook/order-builder/<id>",
            }

        Response (error):
            {ok: False, error: "<code>", message: "<text>",
             login_url: "..."?, state: "<order.state>"?}
        """
        # Anonymous visitors: surface login_required so the OWL
        # component can render a "Sign in to add to quote" CTA
        # instead of a generic AccessError.
        if request.env.user._is_public():
            return {
                "ok": False,
                "error": "login_required",
                "login_url": "/web/signup",
                "message": "Sign in or create a free account to add this "
                           "configuration to your quote.",
            }

        if not session_id:
            return {"ok": False, "error": "missing_session_id"}

        try:
            session = self._authorize_session(int(session_id))
        except ValueError:
            return {"ok": False, "error": "bad_session_id"}
        if isinstance(session, dict):
            return session

        if session.state != "draft":
            return {"ok": False, "error": "session_locked",
                    "message": "This configuration has already been added "
                               "to a quote. Start a new one to add another."}

        # Materialise (or fetch existing) variant for the picks.
        # OCA's create_get_variant calls validate_configuration first;
        # it raises ValidationError on rule violation or missing
        # required values.
        try:
            variant = session.sudo().create_get_variant()
        except (UserError, ValidationError) as exc:
            return {"ok": False, "error": "validation_failed",
                    "message": getattr(exc, "args", [str(exc)])[0]}

        # Resolve or create the user's draft sale.order. Same pattern
        # as southbrook_estimating_website's G9 self-service order
        # creation (one draft per partner; reused across multiple
        # add-to-quote actions until the partner submits it).
        partner = request.env.user.partner_id
        SaleOrder = request.env["sale.order"].sudo()
        if order_id:
            try:
                order = SaleOrder.browse(int(order_id)).exists()
            except ValueError:
                return {"ok": False, "error": "bad_order_id"}
            if not order:
                return {"ok": False, "error": "order_not_found"}
            # Authorize: order must belong to this user's partner.
            if order.partner_id.id != partner.id:
                return {"ok": False, "error": "order_forbidden"}
            if order.state not in ("draft", "sent"):
                return {"ok": False, "error": "order_locked",
                        "state": order.state,
                        "message": f"Cannot add to a {order.state} order; "
                                   f"start a new quote."}
        else:
            order = SaleOrder.search([
                ("partner_id", "=", partner.id),
                ("state", "=", "draft"),
            ], limit=1, order="create_date desc")
            if not order:
                order = SaleOrder.create({"partner_id": partner.id})

        # Add the variant as a new order line. qty=1 by default; the
        # customer can change qty in the Order Builder. Trigger
        # product_id_change so price_unit / name / uom populate from
        # the variant.
        line = request.env["sale.order.line"].sudo().create({
            "order_id": order.id,
            "product_id": variant.id,
            "product_uom_qty": 1,
        })
        if hasattr(line, "product_id_change"):
            line.product_id_change()

        # Lock the session: state='done' + product_id link. A new
        # configuration starts a new session via /state.
        try:
            session.sudo().action_confirm(product_id=variant)
        except Exception:                                   # noqa: BLE001
            # action_confirm failure is non-fatal — the variant was
            # created + the line was added; the session just stays
            # in draft (and the cleanup cron handles it eventually).
            _logger.warning(
                "session.action_confirm failed on session %s; "
                "variant %s + line %s already created.",
                session.id, variant.id, line.id, exc_info=True,
            )

        return {
            "ok": True,
            "variant_id": variant.id,
            "order_id": order.id,
            "order_line_id": line.id,
            "redirect": f"/my/southbrook/order-builder/{order.id}",
        }

    # ------------------------------------------------------------------
    # Shared helper: resolve + authorize a session by id.
    # ------------------------------------------------------------------
    def _authorize_session(self, session_id):
        """Browse + permission-check a session.

        Returns the recordset on success, or an error dict (caller
        bubbles it up as-is to the client) on failure. Sessions can
        only be touched by the user who created them — same shape as
        the customer-flow controller's _southbrook_resolve_order
        authorisation guard.
        """
        Session = request.env["product.config.session"].sudo()
        session = Session.browse(session_id).exists()
        if not session:
            return {"ok": False, "error": "session_not_found"}
        if session.user_id.id != request.env.user.id:
            return {"ok": False, "error": "forbidden",
                    "message": "This configuration session belongs to a "
                               "different user."}
        return session


# =====================================================================
# Phase 4 — bulk product import pipeline.
#
# Lives on a separate controller class because the import endpoints use
# type='http' (file upload + download) rather than type='json'. Keeping
# them on their own class avoids mixing auth modes + makes the
# permission boundary obvious to a reviewer.
# =====================================================================

class SouthbrookImportAPI(http.Controller):
    """Bulk product import endpoints (Phase 4).

    All three routes require an internal user (not portal, not public).
    The bulk-tools bar in the OWL configurator only renders for
    `not user_id.share` — these endpoints enforce the same constraint
    server-side as a defence-in-depth check (so a portal user can't
    POST directly even if they bypass the UI gate).
    """

    # ------------------------------------------------------------------
    # /template — xlsx download with live vocab baked in.
    # ------------------------------------------------------------------
    @http.route(
        "/southbrook/api/import/template",
        type="http",
        auth="user",
        methods=["GET"],
    )
    def import_template(self, **kw):
        """Return an xlsx file matching scripts/gen_import_template.py
        but generated on-the-fly from the live DB vocab.

        Why not just serve the static file from
        ~/Downloads/Southbrook_Product_Import_Template_v1.xlsx?
        Because that snapshot ages — when a new attribute / category /
        UoM lands, the on-disk file's REF_* sheets get stale. Generating
        on-demand keeps the dropdown sources in sync with what the
        validator on /preview will accept.
        """
        if request.env.user.share:
            return request.make_response(
                "Bulk template download is internal-users only.",
                status=403,
            )

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            _logger.error("openpyxl missing in container; can't generate template.")
            return request.make_response(
                "Server is missing openpyxl. Install via "
                "pip install openpyxl in the Odoo container.",
                status=500,
            )

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # PRODUCTS sheet — minimal Phase-4 v1 column set. The full
        # column set (with southbrook_* metadata, BoM, etc.) is shipped
        # by the offline scripts/gen_import_template.py; for the
        # backend template-download we keep this to the v1 importable
        # subset so users aren't tempted to fill columns we'll skip.
        ws = wb.create_sheet("PRODUCTS")
        headers = [
            "default_code", "name", "type", "internal_category", "uom_id",
            "list_price", "standard_price", "sale_ok", "purchase_ok",
            "is_published", "config_ok", "weight",
            "southbrook_category", "southbrook_description",
            "southbrook_dimensions", "southbrook_icon_key",
        ]
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font = Font(color="FFFFFF", bold=True)
            c.fill = PatternFill("solid", fgColor="2F3B52")
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)].width = 18
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        # One example row so the user can see what shapes work.
        example = [
            "SB-DEMO-001", "Demo Cabinet 18\"", "consu", "Goods", "Units",
            295.00, 162.25, "TRUE", "TRUE",
            "FALSE", "TRUE", 14.0,
            "Base", "Demo cabinet for template testing.",
            "18\"W × 34½\"H × 24\"D", "base1",
        ]
        for col_idx, v in enumerate(example, 1):
            ws.cell(row=2, column=col_idx, value=v).fill = (
                PatternFill("solid", fgColor="EEF2FB"))

        # REF_CATEGORIES — live snapshot.
        ref_cats = wb.create_sheet("REF_CATEGORIES")
        ref_cats.cell(row=1, column=1, value="name").font = Font(bold=True)
        for i, cat in enumerate(
                request.env["product.category"].sudo().search([]), 2):
            ref_cats.cell(row=i, column=1, value=cat.name)

        # REF_UOM — live snapshot.
        ref_uom = wb.create_sheet("REF_UOM")
        ref_uom.cell(row=1, column=1, value="name").font = Font(bold=True)
        for i, uom in enumerate(
                request.env["uom.uom"].sudo().search([("active", "=", True)]), 2):
            ref_uom.cell(row=i, column=1, value=uom.name)

        # Stream the workbook to bytes.
        bio = io.BytesIO()
        wb.save(bio)
        data = bio.getvalue()

        return request.make_response(
            data,
            headers=[
                ("Content-Type",
                 "application/vnd.openxmlformats-officedocument."
                 "spreadsheetml.sheet"),
                ("Content-Length", str(len(data))),
                ("Content-Disposition",
                 'attachment; filename="Southbrook_Product_Template.xlsx"'),
            ],
        )

    # ------------------------------------------------------------------
    # /preview — read xlsx, validate, return per-row results. No writes.
    # ------------------------------------------------------------------
    @http.route(
        "/southbrook/api/import/preview",
        type="http",
        auth="user",
        methods=["POST"],
        csrf=False,
    )
    def import_preview(self, **kw):
        """Parse + validate an uploaded xlsx and return a per-row report.

        Multipart form upload — the file comes through the `file` field.
        Response is JSON, even though the route is type='http' (the
        client expects a JSON body it can render in the preview modal).
        """
        result, status = self._import_preview_impl(kw)
        return self._json_response(result, status)

    # ------------------------------------------------------------------
    # /commit — same parse/validate as preview, but writes valid rows.
    # ------------------------------------------------------------------
    @http.route(
        "/southbrook/api/import/commit",
        type="http",
        auth="user",
        methods=["POST"],
        csrf=False,
    )
    def import_commit(self, **kw):
        """Commit valid rows from an uploaded xlsx inside a transaction.

        REQUIRES `confirm` field with value `true` in the form — the
        explicit human-confirmation gate per the Phase 4 stop-point.
        Without it the endpoint refuses (400) rather than silently
        defaulting to "yes".
        """
        result, status = self._import_commit_impl(kw)
        return self._json_response(result, status)

    # ------------------------------------------------------------------
    # Implementation core — pure dict-returning functions. Easier to
    # test in isolation: a test calls these directly and inspects the
    # returned (dict, status_code) tuple without needing to mock
    # request.make_response.
    # ------------------------------------------------------------------
    def _import_preview_impl(self, kw):
        return self._import_run(kw, commit=False)

    def _import_commit_impl(self, kw):
        # Confirm gate runs INSIDE the impl so unit tests of the
        # commit flow can exercise both paths without round-tripping
        # the route wrapper.
        confirm = (kw.get("confirm") or "").lower().strip()
        if confirm != "true":
            return {
                "ok": False, "error": "confirm_required",
                "message": "Commit requires confirm=true in the request body.",
            }, 400
        return self._import_run(kw, commit=True)

    # ------------------------------------------------------------------
    # Shared runner: read xlsx, validate, optionally write.
    # Returns (response_dict, status_code).
    # ------------------------------------------------------------------
    def _import_run(self, kw, commit):
        """Single entry point for preview + commit so the two share
        validation logic byte-for-byte.

        commit=False  → preview pass; nothing written; per-row status
                        for VALID rows is 'preview_ok'.
        commit=True   → writes valid rows inside the request transaction;
                        statuses are 'created' / 'updated' / 'error';
                        invalid rows are 'skipped' with their errors.
        """
        if request.env.user.share:
            return {
                "ok": False, "error": "forbidden",
                "message": "Bulk import is internal-users only.",
            }, 403

        file_storage = request.httprequest.files.get("file")
        if not file_storage:
            return {
                "ok": False, "error": "missing_file",
                "message": "Upload a file under the 'file' form field.",
            }, 400

        try:
            import openpyxl
        except ImportError:
            return {
                "ok": False, "error": "openpyxl_missing",
                "message": "Server is missing openpyxl.",
            }, 500

        # Read + parse the file.
        try:
            wb = openpyxl.load_workbook(file_storage, read_only=True,
                                         data_only=True)
        except Exception as exc:                            # noqa: BLE001
            return {
                "ok": False, "error": "unreadable_file",
                "message": f"Could not parse xlsx: {exc}",
            }, 400

        report = {
            "ok": True,
            "mode": "commit" if commit else "preview",
            "sheets": [],
            "summary": {"valid": 0, "invalid": 0, "skipped_sheets": 0,
                        "created": 0, "updated": 0, "errors": 0},
        }

        for sheet_name in wb.sheetnames:
            if sheet_name == _V1_SHEET:
                self._process_products_sheet(
                    wb[sheet_name], report, commit=commit)
            elif sheet_name in _V2_SHEETS:
                report["sheets"].append({
                    "sheet": sheet_name,
                    "status": "deferred",
                    "message": f"Sheet '{sheet_name}' is recognised but its "
                               f"importer is deferred to v2. Skipping.",
                    "rows": [],
                })
                report["summary"]["skipped_sheets"] += 1
            elif sheet_name in _REF_SHEETS:
                report["sheets"].append({
                    "sheet": sheet_name,
                    "status": "reference",
                    "message": f"Sheet '{sheet_name}' is reference data; "
                               f"importer doesn't read it.",
                    "rows": [],
                })
            else:
                report["sheets"].append({
                    "sheet": sheet_name,
                    "status": "unknown",
                    "message": f"Sheet '{sheet_name}' isn't recognised; "
                               f"importer skipped it.",
                    "rows": [],
                })
                report["summary"]["skipped_sheets"] += 1

        return report, 200

    def _process_products_sheet(self, ws, report, commit):
        """Validate + (optionally) write each row of the PRODUCTS sheet."""
        sheet_report = {
            "sheet": "PRODUCTS",
            "status": "processed",
            "rows": [],
        }

        # Read header row. openpyxl read_only sheets are iter-based so
        # convert the first row to a list of column names.
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else ""
                      for c in next(rows_iter)]
        except StopIteration:
            sheet_report["status"] = "empty"
            report["sheets"].append(sheet_report)
            return

        # Resolve some env shortcuts once.
        Category = request.env["product.category"].sudo()
        Uom = request.env["uom.uom"].sudo()
        Template = request.env["product.template"].sudo()

        # Cache vocab lookups for the validation pass.
        categs_by_name = {c.name: c for c in Category.search([])}
        uoms_by_name = {u.name: u for u in Uom.search([("active", "=", True)])}

        for row_idx, raw in enumerate(rows_iter, start=2):
            row = self._row_to_dict(header, raw)
            # Skip blank rows (every column empty).
            if all((not str(v).strip() if v is not None else True)
                   for v in row.values()):
                continue

            errors = []
            normalised = self._normalise_products_row(row, errors)
            # Vocab lookups
            cat_name = normalised.get("internal_category", "")
            cat = categs_by_name.get(cat_name) if cat_name else None
            if cat_name and not cat:
                errors.append(f"internal_category '{cat_name}' not "
                              f"found in product.category")
            uom_name = normalised.get("uom_id", "") or "Units"
            uom = uoms_by_name.get(uom_name)
            if not uom:
                errors.append(f"uom_id '{uom_name}' not found in uom.uom")

            # Required field check (after normalisation so trimmed values
            # get a fair shake).
            for k in _PRODUCTS_REQUIRED:
                if not normalised.get(k):
                    errors.append(f"{k} is required")

            row_report = {
                "row": row_idx,
                "default_code": normalised.get("default_code") or "",
                "status": None,
                "errors": errors,
            }

            if errors:
                row_report["status"] = (
                    "skipped" if commit else "invalid"
                )
                report["summary"]["invalid"] += 1
                sheet_report["rows"].append(row_report)
                continue

            # Build the write vals. Note: product.template in Odoo 19
            # exposes uom_id but NOT uom_po_id on the template (it's
            # on product.product / removed at the template level).
            vals = {
                "name": normalised["name"],
                "default_code": normalised["default_code"],
                "type": normalised["type"],
                "categ_id": cat.id,
                "uom_id": uom.id,
            }
            for opt in ("list_price", "standard_price", "weight"):
                v = normalised.get(opt)
                if v not in (None, ""):
                    vals[opt] = v
            for opt in ("sale_ok", "purchase_ok", "is_published",
                        "config_ok"):
                v = normalised.get(opt)
                if v is not None:
                    vals[opt] = v
            for opt in ("southbrook_category", "southbrook_description",
                        "southbrook_dimensions", "southbrook_icon_key"):
                v = normalised.get(opt)
                if v not in (None, ""):
                    vals[opt] = v

            if not commit:
                row_report["status"] = "preview_ok"
                row_report["proposed_vals"] = {
                    k: v for k, v in vals.items()
                    if k in ("name", "default_code", "list_price",
                             "southbrook_category", "southbrook_icon_key")
                }
                report["summary"]["valid"] += 1
                sheet_report["rows"].append(row_report)
                continue

            # COMMIT path — upsert by default_code.
            try:
                existing = Template.search(
                    [("default_code", "=", vals["default_code"])], limit=1)
                if existing:
                    existing.write(vals)
                    row_report["status"] = "updated"
                    row_report["product_tmpl_id"] = existing.id
                    report["summary"]["updated"] += 1
                else:
                    new = Template.create(vals)
                    row_report["status"] = "created"
                    row_report["product_tmpl_id"] = new.id
                    report["summary"]["created"] += 1
                report["summary"]["valid"] += 1
            except Exception as exc:                        # noqa: BLE001
                row_report["status"] = "error"
                row_report["errors"].append(
                    f"Write failed: {exc.__class__.__name__}: {exc}")
                report["summary"]["errors"] += 1
                report["summary"]["invalid"] += 1

            sheet_report["rows"].append(row_report)

        report["sheets"].append(sheet_report)

    def _row_to_dict(self, header, raw_tuple):
        """Pair the raw cell values with the column names. Missing
        trailing cells (openpyxl returns shorter tuples for sparse
        rows) get None."""
        out = {}
        for i, name in enumerate(header):
            if not name:
                continue
            try:
                out[name] = raw_tuple[i]
            except IndexError:
                out[name] = None
        return out

    def _normalise_products_row(self, row, errors):
        """Coerce types + validate enums BEFORE the DB lookups + write.

        Returns a new dict with normalised values (strings stripped,
        booleans converted, numerics cast). Anything that fails type
        coercion appends to `errors` and the row gets skipped.
        """
        out = {}

        def _str(k):
            v = row.get(k)
            if v is None: return ""
            return str(v).strip()

        def _num(k):
            v = row.get(k)
            if v is None or (isinstance(v, str) and not v.strip()):
                return None
            try:
                return float(v)
            except (TypeError, ValueError):
                errors.append(f"{k} '{v}' is not numeric")
                return None

        def _bool(k):
            v = row.get(k)
            if v is None: return None
            if isinstance(v, bool): return v
            s = str(v).strip().lower()
            if s in _TRUE_TOKENS: return True
            if s in _FALSE_TOKENS: return False
            errors.append(f"{k} '{v}' is not a boolean")
            return None

        # Strings
        for k in ("default_code", "name", "type", "internal_category",
                  "uom_id", "southbrook_description",
                  "southbrook_dimensions"):
            out[k] = _str(k)

        # Numerics
        for k in ("list_price", "standard_price", "weight"):
            out[k] = _num(k)

        # Booleans
        for k in ("sale_ok", "purchase_ok", "is_published", "config_ok"):
            out[k] = _bool(k)

        # Enum: type must be one of consu/service/combo
        if out["type"] and out["type"] not in ("consu", "service", "combo"):
            errors.append(
                f"type '{out['type']}' must be one of: consu, service, combo")

        # Enum: southbrook_category
        sc = _str("southbrook_category")
        if sc and sc not in _SOUTHBROOK_CATEGORIES:
            errors.append(
                f"southbrook_category '{sc}' must be one of: "
                f"{', '.join(sorted(_SOUTHBROOK_CATEGORIES))}")
        out["southbrook_category"] = sc

        # Enum: southbrook_icon_key
        sik = _str("southbrook_icon_key")
        if sik and sik not in _SOUTHBROOK_ICON_KEYS:
            errors.append(
                f"southbrook_icon_key '{sik}' isn't in the supported set; "
                f"falls back to 'extra' on render but the importer flags it.")
        out["southbrook_icon_key"] = sik

        return out

    def _json_response(self, payload, status=200):
        """Return a JSON body via an http response (since the routes are
        type='http' for the file-upload semantics)."""
        return request.make_response(
            json.dumps(payload),
            status=status,
            headers=[("Content-Type", "application/json")],
        )
