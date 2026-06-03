# SPDX-License-Identifier: LGPL-3.0-only
"""Contract tests for Phase 4 import endpoints.

  GET  /southbrook/api/import/template
  POST /southbrook/api/import/preview
  POST /southbrook/api/import/commit

Unlike the JSON-RPC controllers tested elsewhere, these routes are
type='http' (file upload + download). Rather than stubbing
werkzeug's multipart parser, the tests exercise the validation +
write logic directly via the controller's _process_products_sheet
helper. The HTTP-layer behaviour (Content-Type negotiation,
confirm=true gate) is tested via direct calls to the import_commit
route method with a stubbed request.

Run:

    odoo --no-http --test-enable -u southbrook_configurator_ux \\
        -d <db> --stop-after-init

Or with the explicit tag:

    --test-tags=southbrook_cfg_import
"""
import io
from contextlib import contextmanager
from unittest.mock import MagicMock

import openpyxl

from odoo.tests import TransactionCase, tagged

from odoo.addons.southbrook_configurator_ux.controllers import main as ctrl_main


@contextmanager
def stubbed_request(env, user=None, files=None, form=None):
    """Mock controller.request so the impl methods can be called
    directly from tests.

    The mock surfaces .env, .httprequest.files, and .httprequest.form
    so the importer's file-upload + confirm handling exercise correctly.
    Tests call _import_preview_impl(...) / _import_commit_impl(...)
    which return (result_dict, status_code) tuples — no involvement
    of make_response, so the @http.route wrapper's response-type
    validation isn't a factor.
    """
    saved = ctrl_main.request
    mock = MagicMock()
    mock.env = env if user is None else env(user=user.id)
    mock.session = {}
    mock.params = {}
    mock.httprequest = MagicMock()
    mock.httprequest.files = files or {}
    mock.httprequest.form = form or {}
    ctrl_main.request = mock
    try:
        yield mock
    finally:
        ctrl_main.request = saved


def _build_xlsx(rows, header=None):
    """Helper: build an xlsx in-memory with a PRODUCTS sheet containing
    the given header + rows. Returns a BytesIO ready to feed
    openpyxl.load_workbook."""
    if header is None:
        header = [
            "default_code", "name", "type", "internal_category", "uom_id",
            "list_price", "is_published", "config_ok",
            "southbrook_category", "southbrook_icon_key",
        ]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("PRODUCTS")
    for col_idx, h in enumerate(header, 1):
        ws.cell(row=1, column=col_idx, value=h)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, v in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@tagged("post_install", "-at_install", "southbrook_cfg_import")
class TestImportPipeline(TransactionCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.controller = ctrl_main.SouthbrookImportAPI()
        # Use the admin (internal) user — bulk import requires
        # not user.share.
        cls.admin = cls.env.ref("base.user_admin")
        # Portal user for the forbidden-path tests.
        cls.portal_user = cls.env["res.users"].create({
            "name": "Test Portal Import",
            "login": "test_portal_import_2026@southbrook.test",
            "partner_id": cls.env["res.partner"].create(
                {"name": "Portal Importer"}).id,
            "group_ids": [(6, 0, [cls.env.ref("base.group_portal").id])],
        })

    # ==================================================================
    # Permission gates
    # ==================================================================
    def test_preview_forbidden_for_portal_user(self):
        f = MagicMock()
        with stubbed_request(self.env, user=self.portal_user,
                              files={"file": f}):
            body, status = self.controller._import_preview_impl({})
        self.assertEqual(status, 403)
        self.assertEqual(body["error"], "forbidden")

    def test_commit_requires_confirm_true(self):
        f = MagicMock()
        with stubbed_request(self.env, user=self.admin,
                              files={"file": f}):
            body, status = self.controller._import_commit_impl({})
        self.assertEqual(status, 400)
        self.assertEqual(body["error"], "confirm_required")

    def test_preview_rejects_missing_file(self):
        with stubbed_request(self.env, user=self.admin, files={}):
            body, status = self.controller._import_preview_impl({})
        self.assertEqual(status, 400)
        self.assertEqual(body["error"], "missing_file")

    # ==================================================================
    # Preview happy paths
    # ==================================================================
    def test_preview_valid_rows_marked_preview_ok(self):
        bio = _build_xlsx([
            ["SB-TEST-001", "Test Cabinet", "consu", "Goods", "Units",
             100.0, "TRUE", "TRUE", "Base", "base1"],
            ["SB-TEST-002", "Other Cabinet", "consu", "Goods", "Units",
             200.0, "TRUE", "TRUE", "Wall", "wall1"],
        ])
        with stubbed_request(self.env, user=self.admin,
                              files={"file": bio}):
            body, status = self.controller._import_preview_impl({})
        self.assertEqual(status, 200)
        self.assertTrue(body["ok"])
        self.assertEqual(body["mode"], "preview")
        self.assertEqual(body["summary"]["valid"], 2)
        self.assertEqual(body["summary"]["invalid"], 0)
        for sheet in body["sheets"]:
            if sheet["sheet"] == "PRODUCTS":
                for row in sheet["rows"]:
                    self.assertEqual(row["status"], "preview_ok")
                break

    def test_preview_invalid_rows_marked_invalid_with_errors(self):
        bio = _build_xlsx([
            # Missing default_code
            ["", "Nameless", "consu", "Goods", "Units",
             100.0, "TRUE", "TRUE", "Base", "base1"],
            # Bad type enum
            ["SB-BAD-002", "Bad Type", "rocket", "Goods", "Units",
             50.0, "TRUE", "TRUE", "Base", "base1"],
            # Unknown category
            ["SB-BAD-003", "Bad Cat", "consu", "Mythical", "Units",
             50.0, "TRUE", "TRUE", "Base", "base1"],
            # Bad southbrook_category enum
            ["SB-BAD-004", "Bad SCat", "consu", "Goods", "Units",
             50.0, "TRUE", "TRUE", "Imaginary", "base1"],
        ])
        with stubbed_request(self.env, user=self.admin,
                              files={"file": bio}):
            body, status = self.controller._import_preview_impl({})
        self.assertEqual(body["summary"]["invalid"], 4)
        self.assertEqual(body["summary"]["valid"], 0)
        for sheet in body["sheets"]:
            if sheet["sheet"] == "PRODUCTS":
                for row in sheet["rows"]:
                    self.assertEqual(row["status"], "invalid")
                    self.assertGreater(len(row["errors"]), 0)
                break

    def test_preview_v2_sheets_marked_deferred(self):
        # Build a workbook with a PRODUCTS sheet AND an
        # ATTRIBUTE_LINES sheet — the latter should be flagged as
        # deferred to v2, not processed.
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("PRODUCTS")
        for i, h in enumerate(["default_code", "name", "type",
                                "internal_category", "uom_id"], 1):
            ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=1, value="SB-V1-001")
        ws.cell(row=2, column=2, value="V1 Test")
        ws.cell(row=2, column=3, value="consu")
        ws.cell(row=2, column=4, value="Goods")
        ws.cell(row=2, column=5, value="Units")
        wb.create_sheet("ATTRIBUTE_LINES")  # empty; just there to be detected
        bio = io.BytesIO()
        wb.save(bio); bio.seek(0)
        with stubbed_request(self.env, user=self.admin,
                              files={"file": bio}):
            body, status = self.controller._import_preview_impl({})
        statuses = {s["sheet"]: s["status"] for s in body["sheets"]}
        self.assertEqual(statuses["PRODUCTS"], "processed")
        self.assertEqual(statuses["ATTRIBUTE_LINES"], "deferred")
        self.assertEqual(body["summary"]["skipped_sheets"], 1)

    # ==================================================================
    # Commit happy paths
    # ==================================================================
    def test_commit_creates_new_product(self):
        bio = _build_xlsx([
            ["SB-IMP-001", "Imported Cabinet", "consu", "Goods", "Units",
             150.0, "TRUE", "TRUE", "Base", "base1"],
        ])
        with stubbed_request(self.env, user=self.admin,
                              files={"file": bio}):
            body, status = self.controller._import_commit_impl(
                {"confirm": "true"})
        self.assertEqual(status, 200)
        self.assertTrue(body["ok"])
        self.assertEqual(body["summary"]["created"], 1)
        # Verify in DB
        created = self.env["product.template"].search(
            [("default_code", "=", "SB-IMP-001")])
        self.assertTrue(created)
        self.assertEqual(created.name, "Imported Cabinet")
        self.assertEqual(created.southbrook_category, "Base")
        self.assertEqual(created.southbrook_icon_key, "base1")

    def test_commit_updates_existing_product(self):
        # Pre-create a product to test upsert (update path).
        Cat = self.env["product.category"]
        goods = Cat.search([("name", "=", "Goods")], limit=1) or Cat.create(
            {"name": "Goods"})
        Uom = self.env["uom.uom"]
        units = Uom.search([("name", "=", "Units"), ("active", "=", True)],
                            limit=1)
        pre = self.env["product.template"].create({
            "name": "Old Name",
            "default_code": "SB-UPS-001",
            "type": "consu",
            "categ_id": goods.id,
            "uom_id": units.id,
            "list_price": 99.0,
            "southbrook_category": "Wall",
        })
        bio = _build_xlsx([
            ["SB-UPS-001", "Renamed Cabinet", "consu", "Goods", "Units",
             175.0, "TRUE", "TRUE", "Base", "base1"],
        ])
        with stubbed_request(self.env, user=self.admin,
                              files={"file": bio}):
            body, status = self.controller._import_commit_impl(
                {"confirm": "true"})
        self.assertEqual(body["summary"]["updated"], 1)
        self.assertEqual(body["summary"]["created"], 0)
        pre.invalidate_recordset()
        self.assertEqual(pre.name, "Renamed Cabinet")
        self.assertEqual(pre.list_price, 175.0)
        self.assertEqual(pre.southbrook_category, "Base")

    def test_commit_skips_invalid_rows_keeps_valid(self):
        bio = _build_xlsx([
            ["SB-MIX-001", "Good 1", "consu", "Goods", "Units",
             50.0, "TRUE", "TRUE", "Base", "base1"],
            # Bad: missing name
            ["SB-MIX-002", "", "consu", "Goods", "Units",
             50.0, "TRUE", "TRUE", "Base", "base1"],
            ["SB-MIX-003", "Good 3", "consu", "Goods", "Units",
             50.0, "TRUE", "TRUE", "Wall", "wall1"],
        ])
        with stubbed_request(self.env, user=self.admin,
                              files={"file": bio}):
            body, status = self.controller._import_commit_impl(
                {"confirm": "true"})
        self.assertEqual(body["summary"]["created"], 2)
        self.assertEqual(body["summary"]["invalid"], 1)
        self.assertTrue(self.env["product.template"].search(
            [("default_code", "=", "SB-MIX-001")]))
        self.assertFalse(self.env["product.template"].search(
            [("default_code", "=", "SB-MIX-002")]))
        self.assertTrue(self.env["product.template"].search(
            [("default_code", "=", "SB-MIX-003")]))
