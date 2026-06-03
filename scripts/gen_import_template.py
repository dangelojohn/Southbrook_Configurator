#!/usr/bin/env python3
# SPDX-License-Identifier: LGPL-3.0-only
"""Generate the Southbrook master product-import xlsx template.

The output workbook drives the Phase-4 import endpoint of the
southbrook_configurator_ux module. One row per product.template; linked
sheets for attribute_lines, attribute_values, BoM headers + lines,
hardware BoM, and accessory options. Read-only REF_* sheets ship the
live attribute / category / UoM vocabulary as both dropdown sources
and as a snapshot for auditing.

USAGE
-----

  # Refresh vocab from the live QNAP container, then build:
  scripts/gen_import_template.py --refresh

  # Build from the last cached vocab in /tmp (faster, no network):
  scripts/gen_import_template.py

  # Build to a different path:
  scripts/gen_import_template.py --out ~/Desktop/template.xlsx

DEPENDENCIES
------------

  pip install openpyxl

OUTPUT (default)
----------------

  ~/Downloads/Southbrook_Product_Import_Template_v1.xlsx

WHEN TO RE-RUN
--------------

  - After new attributes / attribute values land in Odoo
  - After new product.category records are added
  - After the Southbrook_Cabinet_Engineering_System.xlsx column shape
    changes (we mirror its ERP EXPORT (ODOO) table structure)
  - Before a Phase-4 importer release so reviewers can see the
    template shape against fresh DB state

Importer endpoint contract this template ships against:

  POST /southbrook/api/import/preview   { file: <multipart xlsx> }
        → returns per-row {sheet, row, status, errors[]}
  POST /southbrook/api/import/commit    { file: <multipart xlsx>,
                                          confirm: true,
                                          dry_run: false }
        → 400 without confirm:true; writes inside one transaction;
          returns per-row commit log + summary counters.
"""
from __future__ import annotations

import argparse
import datetime
import os
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
except ImportError:
    sys.stderr.write(
        "openpyxl missing. Install with: pip install openpyxl\n"
    )
    sys.exit(1)


# =====================================================================
# Configuration
# =====================================================================

VERSION = "1.0.0"

DEFAULT_OUT = Path.home() / "Downloads" / "Southbrook_Product_Import_Template_v1.xlsx"

# QNAP container access (read-only vocab pull; safe to run anytime).
SSH_HOST = "admin@192.168.68.108"
DOCKER_BIN = "/share/CACHEDEV3_DATA/.qpkg/container-station/bin/docker"
DOCKER_SOCK = "unix:///var/run/system-docker.sock"
PG_CTR = "southbrook-postgres"
PG_DB = "southbrook"
PG_USER = "odoo"

# Cached snapshots — refreshed by --refresh, read on every build.
CACHE = Path("/tmp")
ATTRS_TSV = CACHE / "sb_attrs.tsv"
CATEGS_TSV = CACHE / "sb_categs.tsv"
UOM_TSV = CACHE / "sb_uom.tsv"
CABINETS_TSV = CACHE / "sb_cabinets.tsv"


# =====================================================================
# Vocab pulling
# =====================================================================

def _psql(sql: str) -> str:
    """Run a psql query against the southbrook DB via SSH+docker; return tab-separated stdout."""
    cmd = (
        f"{DOCKER_BIN} -H {DOCKER_SOCK} exec {PG_CTR} "
        f"psql -U {PG_USER} {PG_DB} -tA -F$'\\t' -c \"{sql}\""
    )
    try:
        result = subprocess.run(
            ["ssh", "-o", "BatchMode=yes", SSH_HOST, cmd],
            check=True, capture_output=True, text=True, timeout=30,
        )
    except subprocess.CalledProcessError as exc:
        sys.stderr.write(
            f"\nSSH/psql query failed:\n  {sql}\n  stderr: {exc.stderr}\n"
            f"  (Is the QNAP reachable? Is the container running? Try:\n"
            f"   ssh {SSH_HOST} '{DOCKER_BIN} -H {DOCKER_SOCK} ps')\n\n"
        )
        sys.exit(2)
    except subprocess.TimeoutExpired:
        sys.stderr.write("psql query timed out after 30s — QNAP unreachable?\n")
        sys.exit(2)
    return result.stdout


def refresh_vocab():
    """Re-pull attribute / category / UoM / cabinet snapshots from the live DB.

    Writes to /tmp/sb_*.tsv. The build step reads these.
    """
    sys.stderr.write("Refreshing vocab from QNAP …\n")

    ATTRS_TSV.write_text(_psql(
        "SELECT pa.name->>'en_US', pa.display_type, pa.create_variant, "
        "       string_agg(pav.name->>'en_US', '|' ORDER BY pav.sequence) "
        "FROM product_attribute pa "
        "LEFT JOIN product_attribute_value pav ON pav.attribute_id = pa.id "
        "WHERE pa.name->>'en_US' NOT IN ('Custom') "
        "GROUP BY pa.id, pa.name, pa.display_type, pa.create_variant "
        "ORDER BY pa.sequence, pa.id;"
    ))
    sys.stderr.write(f"  · {ATTRS_TSV}  ({sum(1 for _ in ATTRS_TSV.open())} rows)\n")

    CATEGS_TSV.write_text(_psql(
        "SELECT id, name FROM product_category ORDER BY parent_id NULLS FIRST, id;"
    ))
    sys.stderr.write(f"  · {CATEGS_TSV}  ({sum(1 for _ in CATEGS_TSV.open())} rows)\n")

    UOM_TSV.write_text(_psql(
        "SELECT id, name->>'en_US' FROM uom_uom WHERE active = true ORDER BY id LIMIT 30;"
    ))
    sys.stderr.write(f"  · {UOM_TSV}  ({sum(1 for _ in UOM_TSV.open())} rows)\n")

    # Cabinets join on ir_model_data so we hit the Q8 xml_ids even when
    # template.default_code has blanked (a known v19 behaviour for
    # dynamic-variant templates with >0 materialised variants).
    CABINETS_TSV.write_text(_psql(
        "SELECT imd.name AS xml_id, pt.id, COALESCE(pt.default_code, ''), "
        "       pt.name->>'en_US', pt.list_price, "
        "       COALESCE(pt.southbrook_category, ''), "
        "       COALESCE(pt.southbrook_dimensions, ''), "
        "       COALESCE(pt.southbrook_icon_key, ''), "
        "       COALESCE(pt.southbrook_description->>'en_US', '') "
        "FROM product_template pt "
        "JOIN ir_model_data imd ON imd.res_id = pt.id "
        "  AND imd.model = 'product.template' "
        "WHERE imd.module = 'southbrook_estimating' "
        "ORDER BY imd.name;"
    ))
    sys.stderr.write(f"  · {CABINETS_TSV}  ({sum(1 for _ in CABINETS_TSV.open())} rows)\n")
    sys.stderr.write("Done.\n")


def load_tsv(path: Path) -> list[list[str]]:
    if not path.exists():
        sys.stderr.write(
            f"\nMissing vocab cache: {path}\n"
            f"Run with --refresh to pull from the QNAP container.\n\n"
        )
        sys.exit(3)
    rows = []
    for line in path.read_text().splitlines():
        if not line:
            continue
        rows.append(line.split("\t"))
    return rows


# =====================================================================
# Styles
# =====================================================================

HEADER_FILL = PatternFill("solid", fgColor="2F3B52")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
INPUT_FILL = PatternFill("solid", fgColor="EEF2FB")
REF_FILL = PatternFill("solid", fgColor="FFF7E8")
NOTE_FONT = Font(italic=True, color="6B7488", size=9)
TITLE_FONT = Font(bold=True, size=14, color="2F3B52")
SUB_FONT = Font(bold=True, size=11, color="2F3B52")
THIN = Side(border_style="thin", color="D6DAE3")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_header(ws, headers, row=1):
    for col_idx, (name, comment) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
        if comment:
            cell.comment = Comment(comment, "Southbrook Importer")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(name) + 2)
    ws.row_dimensions[row].height = 38
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# =====================================================================
# Column definitions  (sheet name → list of (header, tooltip))
# =====================================================================

PRODUCT_COLS = [
    ("external_id",            "Optional. Stable xml_id like 'southbrook_estimating.base_1dr'. If blank, importer generates one from default_code."),
    ("default_code",           "REQUIRED. SKU. Primary key for upsert. Must be unique."),
    ("name",                   "REQUIRED. Display name."),
    ("type",                   "REQUIRED. 'consu' for consumable, 'service' for service. Cabinets typically 'consu' (Odoo 19 replaces 'product' with 'consu'+is_storable=TRUE)."),
    ("is_storable",            "TRUE for stockable cabinets/parts. FALSE for services."),
    ("internal_category",      "REQUIRED. Internal category from REF_CATEGORIES. e.g. 'Goods', 'Fasteners', 'Finishing'."),
    ("uom_id",                 "REQUIRED. Unit of measure. Usually 'Units'. See REF_UOM."),
    ("list_price",             "REQUIRED for sale_ok products. Sale price in CAD."),
    ("standard_price",         "Cost in CAD. Defaults to list_price * 0.55 if blank (placeholder)."),
    ("sale_ok",                "TRUE if sellable. Default TRUE."),
    ("purchase_ok",            "TRUE if purchasable. Default TRUE."),
    ("active",                 "TRUE to enable on install. Default TRUE."),
    ("is_published",           "TRUE to publish on website /shop/<slug>. Default FALSE for backend-only products."),
    ("config_ok",              "TRUE if this product is configurable via OCA product_configurator. Cabinets = TRUE."),
    ("weight",                 "Kg. Optional."),
    ("volume",                 "m^3. Optional."),
    ("description_sale",       "Internal sales-rep notes (Odoo description_sale field)."),
    ("southbrook_category",    "REQUIRED for cabinets. One of: Wall / Base / Drawer / Tall / Vanity / Extras. Drives the catalog picker pill."),
    ("southbrook_description", "One-sentence customer-facing description shown on the catalog card. Translatable."),
    ("southbrook_dimensions",  "Reference dimensions for the catalog card. e.g. '18\"W x 34 1/2\"H x 24\"D' or 'Varies'."),
    ("southbrook_icon_key",    "Icon key for the catalog card SVG. One of: wall1, wall2, base1, base2, drawer, sink, pantry, oven, corner, vanity, extra, worktop. Unknown -> 'extra'."),
    ("manufacture_route",      "TRUE if produced internally (assigns Manufacture route). FALSE for purchased components."),
    ("default_code_alias",     "Optional alternative SKU codes (semicolon-separated). Importer matches on these for upsert too."),
]

ATTR_LINE_COLS = [
    ("parent_sku",     "REQUIRED. SKU from PRODUCTS sheet."),
    ("attribute_name", "REQUIRED. Attribute name. See REF_ATTRIBUTES for valid options."),
    ("value_names",    "REQUIRED. Pipe-separated list of value names to expose. e.g. 'Wall' or 'LH (Left Hand)|RH (Right Hand)'. Must match REF_ATTRIBUTES exactly."),
    ("sequence",       "Optional. Display order within the configurator. Defaults to attribute.sequence."),
    ("required",       "TRUE if customer must pick a value. Defaults to FALSE."),
]

ATTR_VAL_COLS = [
    ("parent_sku",     "REQUIRED. SKU from PRODUCTS sheet."),
    ("attribute_name", "REQUIRED. Attribute (matches an ATTRIBUTE_LINES row for this SKU)."),
    ("value_name",     "REQUIRED. Specific value (matches one in the ATTRIBUTE_LINES.value_names list)."),
    ("price_extra",    "Optional. Surcharge in CAD applied when this value is chosen. Default 0."),
    ("html_color",     "Optional. For 'color' display_type attributes (e.g. Finish). HEX like '#5a3b28'."),
    ("excluded_for",   "Optional. Comma-separated list of attribute_name=value pairs that exclude this value. e.g. 'Box Material=White Melamine'."),
]

BOM_HEAD_COLS = [
    ("parent_sku",       "REQUIRED. SKU of the manufactured product (from PRODUCTS sheet)."),
    ("bom_code",         "Optional. Human-readable BoM ref. Defaults to '<parent_sku> BOM'."),
    ("bom_type",         "REQUIRED. 'normal' for assembly, 'phantom' for kit, 'subcontract' for outsourced."),
    ("product_qty",      "REQUIRED. Quantity produced per BoM. Default 1."),
    ("product_uom",      "REQUIRED. UoM of produced quantity. Usually 'Units'."),
    ("routing_id",       "Optional. mrp.routing reference. Blank uses the default workcenter."),
    ("ready_to_produce", "Optional. 'asap' or 'all'. Default 'asap'."),
]

BOM_LINE_COLS = [
    ("parent_sku",     "REQUIRED. SKU of the manufactured product (matches BOM_HEADERS.parent_sku)."),
    ("component_sku",  "REQUIRED. SKU of the component (must exist on PRODUCTS sheet OR be a pre-existing product). e.g. 'SB-BASE-1DR-GBL'."),
    ("component_name", "Display name of the component. e.g. 'Gable / Side', 'Door', 'Adjustable Shelf'."),
    ("product_qty",    "REQUIRED. Quantity per parent. Numeric."),
    ("product_uom",    "REQUIRED. UoM. Usually 'Units'."),
    ("length_mm",      "Cut length in mm. For panels / extrusions."),
    ("width_mm",       "Cut width in mm."),
    ("thickness_mm",   "Material thickness in mm."),
    ("material",       "Material code. e.g. '18mm Carcass', '18mm Door/Front', '6mm Back'."),
    ("grain",          "Grain direction: 'Length', 'Width', 'None'. Drives CNC nest rotation."),
    ("rotate_ok",      "TRUE if part can rotate 90deg during nesting. Default TRUE for 'None'."),
    ("edge_bands",     "Comma-separated edges to band: e.g. 'F,B,L,R' for all four."),
    ("seq",            "Optional. Sequence within the BoM (for display)."),
    ("notes",          "Optional free text."),
]

HW_COLS = [
    ("parent_sku",      "REQUIRED. SKU of the cabinet this hardware is for."),
    ("component_type",  "REQUIRED. e.g. 'Hinge', 'Drawer Slide', 'Leg', 'Leg Clip', 'Toe-Kick Clip', 'Shelf Pin', 'Handle'."),
    ("vendor_sku",      "Vendor SKU. e.g. 'BLUM-71B3550' or 'HET-9243625'."),
    ("vendor_name",     "Vendor name. e.g. 'Blum Canada', 'Hettich', 'Richelieu'."),
    ("description",     "Description."),
    ("qty_per_cabinet", "REQUIRED. Qty per parent cabinet."),
    ("uom",             "UoM. Usually 'Units'."),
    ("unit_cost",       "Cost in CAD per unit. Optional."),
    ("load_class",      "For slides/hinges: 'Light 35kg', 'Medium 45kg', 'Heavy 60kg'."),
    ("notes",           "Free text."),
]

ACC_COLS = [
    ("parent_sku",       "REQUIRED. SKU this accessory option applies to."),
    ("accessory_sku",    "REQUIRED. SKU of the accessory product (must exist on PRODUCTS or be created here)."),
    ("accessory_type",   "REQUIRED. One of: End Panel, Filler, Cornice, Pelmet, Plinth, Internal Pull-Out, Soft-Close, Drawer Organisers."),
    ("display_name",     "Display name in the configurator's Accessories multi-select."),
    ("price_extra",      "Surcharge in CAD when selected. Defaults to accessory_sku.list_price."),
    ("default_selected", "TRUE if pre-selected in the configurator. Default FALSE."),
    ("notes",            ""),
]


# =====================================================================
# Worked-example seed data
# =====================================================================

EXAMPLE_ATTR_LINES = [
    ("SB-BASE-1DR", "Family",         "Base"),
    ("SB-BASE-1DR", "Width",          "9 in|12 in|15 in|18 in|21 in"),
    ("SB-BASE-1DR", "Series",         "Contractor Series|Contemporary|Elegance|Signature"),
    ("SB-BASE-1DR", "Box Material",   "White Melamine|Maple"),
    ("SB-BASE-1DR", "Door Style",     "Thermofoil Slab — White|Five-Piece Woodgrain|Custom (Signature)"),
    ("SB-BASE-1DR", "Finish",         "White|Maple Stain|Cherry Stain|Walnut Stain|Custom"),
    ("SB-BASE-1DR", "Hinge Side",     "LH (Left Hand)|RH (Right Hand)"),
    ("SB-BASE-1DR", "Finished Sides", "None|Left|Right|Both"),
    ("SB-BASE-1DR", "Gables",         "Standard|Finished|Decorative"),
    ("SB-BASE-1DR", "Handle",         "Bar Pull|Knob|Cup Pull|Integrated|None"),
    ("SB-BASE-1DR", "Door Count",     "1"),
]

EXAMPLE_ATTR_VALUES = [
    ("SB-BASE-1DR", "Width",        "21 in",                  60.0,  "",        ""),
    ("SB-BASE-1DR", "Width",        "18 in",                  30.0,  "",        ""),
    ("SB-BASE-1DR", "Series",       "Signature",              160.0, "",        ""),
    ("SB-BASE-1DR", "Box Material", "Maple",                  55.0,  "",        ""),
    ("SB-BASE-1DR", "Door Style",   "Five-Piece Woodgrain",   45.0,  "",        ""),
    ("SB-BASE-1DR", "Door Style",   "Custom (Signature)",     120.0, "",        "Series=Contractor Series,Series=Contemporary,Series=Elegance"),
    ("SB-BASE-1DR", "Finish",       "Maple Stain",            20.0,  "#d9a566", "Box Material=White Melamine"),
    ("SB-BASE-1DR", "Finish",       "Cherry Stain",           25.0,  "#8a3b2a", "Box Material=White Melamine"),
    ("SB-BASE-1DR", "Finish",       "Walnut Stain",           30.0,  "#5a3b28", "Box Material=White Melamine"),
    ("SB-BASE-1DR", "Finish",       "Custom",                 75.0,  "#b9a07a", "Series=Contractor Series,Series=Contemporary,Series=Elegance"),
]

EXAMPLE_BOM_LINES = [
    ("SB-BASE-1DR", "SB-BASE-1DR-GBL", "Gable / Side",     2, "Units", 720, 560, 18, "18mm Carcass",  "Length", "FALSE", "F",       1),
    ("SB-BASE-1DR", "SB-BASE-1DR-BOT", "Bottom",           1, "Units", 564, 560, 18, "18mm Carcass",  "Length", "TRUE",  "F",       2),
    ("SB-BASE-1DR", "SB-BASE-1DR-TOP", "Top (stretcher)",  2, "Units", 564, 100, 18, "18mm Carcass",  "Length", "TRUE",  "F",       3),
    ("SB-BASE-1DR", "SB-BASE-1DR-SHF", "Adjustable Shelf", 1, "Units", 560, 540, 18, "18mm Carcass",  "Length", "TRUE",  "F",       4),
    ("SB-BASE-1DR", "SB-BASE-1DR-BCK", "Back Panel",       1, "Units", 700, 580, 6,  "6mm Back",      "Length", "TRUE",  "",        5),
    ("SB-BASE-1DR", "SB-BASE-1DR-TKB", "Toe Kick Board",   1, "Units", 592, 100, 18, "18mm Carcass",  "Length", "FALSE", "F",       6),
    ("SB-BASE-1DR", "SB-BASE-1DR-DOR", "Door",             1, "Units", 716, 596, 18, "18mm Door",     "Length", "FALSE", "F,B,L,R", 7),
]

EXAMPLE_HW = [
    ("SB-BASE-1DR", "Hinge",        "BLUM-71B3550",    "Blum Canada", "Full-overlay soft-close hinge",  2, "Units", 6.80,  "",            ""),
    ("SB-BASE-1DR", "Leg",          "BLUM-LEG-100",    "Blum Canada", "Adjustable plastic leg 100mm",   4, "Units", 1.20,  "",            ""),
    ("SB-BASE-1DR", "Leg Clip",     "BLUM-CLIP-101",   "Blum Canada", "Leg clip / toe-kick mount",      4, "Units", 0.55,  "",            ""),
    ("SB-BASE-1DR", "Shelf Pin",    "RIC-5MM-PIN",     "Richelieu",   "5mm shelf pin",                  4, "Units", 0.08,  "",            ""),
    ("SB-BASE-1DR", "Handle",       "RIC-BAR-128",     "Richelieu",   "Bar pull 128mm",                 1, "Units", 4.50,  "",            "Configurator-overridable"),
    ("SB-BASE-2DR", "Hinge",        "BLUM-71B3550",    "Blum Canada", "Full-overlay soft-close hinge",  4, "Units", 6.80,  "",            ""),
    ("SB-DRAWER",   "Drawer Slide", "BLUM-TANDEM-500", "Blum Canada", "Tandem undermount soft-close",   3, "Pair",  18.00, "Medium 45kg", ""),
]

EXAMPLE_ACC = [
    ("SB-BASE-1DR", "SB-ACC-SOFT-CLOSE",  "Soft-Close",        "Soft-Close upgrade",         24.0, "FALSE", ""),
    ("SB-BASE-1DR", "SB-ACC-PULLOUT",     "Internal Pull-Out", "Internal detergent pull-out", 65.0, "FALSE", ""),
    ("SB-BASE-1DR", "SB-ACC-END-PANEL-L", "End Panel",         "Left end panel (visible)",  120.0, "FALSE", "Phase 2 - gated on Finished Sides"),
    ("SB-DRAWER",   "SB-ACC-ORGANISER",   "Drawer Organisers", "Cutlery organiser inserts",  38.0, "FALSE", ""),
]

# Column -> Odoo field crosswalk for REF_FIELDS
FIELD_MAP = [
    ("PRODUCTS", "external_id",            "ir.model.data", "name",                    "create-or-update by name",          "Phase 1"),
    ("PRODUCTS", "default_code",           "product.template", "default_code",         "direct + upsert key",               "Phase 1"),
    ("PRODUCTS", "name",                   "product.template", "name",                 "direct (translatable)",             "Phase 1"),
    ("PRODUCTS", "type",                   "product.template", "type",                 "direct",                            "Phase 1"),
    ("PRODUCTS", "is_storable",            "product.template", "is_storable",          "direct",                            "Phase 1"),
    ("PRODUCTS", "internal_category",      "product.template", "categ_id",             "name lookup on REF_CATEGORIES",     "Phase 1"),
    ("PRODUCTS", "uom_id",                 "product.template", "uom_id",               "name lookup on REF_UOM",            "Phase 1"),
    ("PRODUCTS", "list_price",             "product.template", "list_price",           "direct",                            "Phase 1"),
    ("PRODUCTS", "standard_price",         "product.template", "standard_price",       "direct",                            "Phase 1"),
    ("PRODUCTS", "sale_ok",                "product.template", "sale_ok",              "boolean",                           "Phase 1"),
    ("PRODUCTS", "purchase_ok",            "product.template", "purchase_ok",          "boolean",                           "Phase 1"),
    ("PRODUCTS", "active",                 "product.template", "active",               "boolean",                           "Phase 1"),
    ("PRODUCTS", "is_published",           "product.template", "is_published",         "boolean",                           "Phase 1"),
    ("PRODUCTS", "config_ok",              "product.template", "config_ok",            "boolean",                           "Phase 1"),
    ("PRODUCTS", "weight",                 "product.template", "weight",               "numeric",                           "Phase 1"),
    ("PRODUCTS", "volume",                 "product.template", "volume",               "numeric",                           "Phase 1"),
    ("PRODUCTS", "description_sale",       "product.template", "description_sale",     "translatable text",                 "Phase 1"),
    ("PRODUCTS", "southbrook_category",    "product.template", "southbrook_category",  "Selection",                         "Phase 1 (added 2026-06-02)"),
    ("PRODUCTS", "southbrook_description", "product.template", "southbrook_description","Char translatable",                "Phase 1 (added 2026-06-02)"),
    ("PRODUCTS", "southbrook_dimensions",  "product.template", "southbrook_dimensions","Char",                              "Phase 1 (added 2026-06-02)"),
    ("PRODUCTS", "southbrook_icon_key",    "product.template", "southbrook_icon_key",  "Char",                              "Phase 1 (added 2026-06-02)"),
    ("PRODUCTS", "manufacture_route",      "product.template", "route_ids",            "name lookup + (4,id)",              "Phase 1"),
    ("ATTRIBUTE_LINES",  "parent_sku",     "product.template", "-",                    "join key for upsert",               "Phase 1"),
    ("ATTRIBUTE_LINES",  "attribute_name", "product.template.attribute.line", "attribute_id", "name lookup on product.attribute", "Phase 1"),
    ("ATTRIBUTE_LINES",  "value_names",    "product.template.attribute.line", "value_ids",    "name lookup (pipe-separated)",     "Phase 1"),
    ("ATTRIBUTE_LINES",  "sequence",       "product.template.attribute.line", "sequence",     "integer",                          "Phase 1"),
    ("ATTRIBUTE_VALUES", "parent_sku",     "product.template", "-",                    "join",                              "Phase 1"),
    ("ATTRIBUTE_VALUES", "attribute_name", "product.template.attribute.value", "attribute_id",            "name lookup",   "Phase 1"),
    ("ATTRIBUTE_VALUES", "value_name",     "product.template.attribute.value", "product_attribute_value_id", "name lookup","Phase 1"),
    ("ATTRIBUTE_VALUES", "price_extra",    "product.template.attribute.value", "price_extra",             "numeric",       "Phase 1"),
    ("ATTRIBUTE_VALUES", "html_color",     "product.attribute.value",          "html_color",              "hex string",    "Phase 1"),
    ("ATTRIBUTE_VALUES", "excluded_for",   "product.config.line",              "-",                       "Phase 2 - wires to rule engine", "Phase 2"),
    ("BOM_HEADERS", "parent_sku",          "mrp.bom",          "product_tmpl_id",      "SKU lookup on product.template",   "Phase 1"),
    ("BOM_HEADERS", "bom_code",            "mrp.bom",          "code",                 "direct",                           "Phase 1"),
    ("BOM_HEADERS", "bom_type",            "mrp.bom",          "type",                 "selection",                        "Phase 1"),
    ("BOM_HEADERS", "product_qty",         "mrp.bom",          "product_qty",          "numeric",                          "Phase 1"),
    ("BOM_HEADERS", "product_uom",         "mrp.bom",          "product_uom_id",       "name lookup on REF_UOM",           "Phase 1"),
    ("BOM_LINES", "parent_sku",            "mrp.bom",          "-",                    "join",                             "Phase 1"),
    ("BOM_LINES", "component_sku",         "mrp.bom.line",     "product_id",           "SKU lookup on product.product (creates if missing)", "Phase 1"),
    ("BOM_LINES", "product_qty",           "mrp.bom.line",     "product_qty",          "numeric",                          "Phase 1"),
    ("BOM_LINES", "length_mm",             "mrp.bom.line",     "x_length_mm",          "Custom Southbrook field",          "Phase 2 (new field)"),
    ("BOM_LINES", "width_mm",              "mrp.bom.line",     "x_width_mm",           "Custom Southbrook field",          "Phase 2 (new field)"),
    ("BOM_LINES", "thickness_mm",          "mrp.bom.line",     "x_thickness_mm",       "Custom Southbrook field",          "Phase 2 (new field)"),
    ("BOM_LINES", "material",              "mrp.bom.line",     "x_material",           "Custom Southbrook field",          "Phase 2 (new field)"),
    ("BOM_LINES", "grain",                 "mrp.bom.line",     "x_grain",              "Custom Southbrook field",          "Phase 2 (new field)"),
    ("BOM_LINES", "edge_bands",            "mrp.bom.line",     "x_edge_bands",         "Custom Southbrook field",          "Phase 2 (new field)"),
    ("HARDWARE_BOM", "parent_sku",         "mrp.bom",          "-",                    "join",                             "Phase 1"),
    ("HARDWARE_BOM", "component_type",     "product.template", "categ_id",             "auto-routes to 'Fasteners' / 'Hinges' etc.", "Phase 1"),
    ("HARDWARE_BOM", "vendor_sku",         "product.template", "default_code",         "creates new product if SKU absent","Phase 1"),
    ("HARDWARE_BOM", "vendor_name",        "res.partner",      "name",                 "name lookup on res.partner where supplier_rank > 0", "Phase 1"),
    ("ACCESSORIES", "parent_sku",          "product.template", "-",                    "join",                             "Phase 1"),
    ("ACCESSORIES", "accessory_sku",       "product.template", "default_code",         "creates new product if SKU absent","Phase 1"),
    ("ACCESSORIES", "accessory_type",      "product.attribute.value", "name",          "name lookup on Accessory Type attribute", "Phase 1"),
]


# =====================================================================
# Build the workbook
# =====================================================================

def build(out: Path, attrs, categs, uoms, cabinets):
    generated_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # -------- Instructions --------
    ws = wb.create_sheet("Instructions", 0)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90
    intro = [
        ("SOUTHBROOK CABINETRY — Master Product Import Template", TITLE_FONT),
        (f"Template version: {VERSION}    Generated: {generated_at}", NOTE_FONT),
        ("", None),
        ("PURPOSE", SUB_FONT),
        (None,
         "One spreadsheet -> one Odoo 19 CE product. Fill this in and the\n"
         "Phase-4 importer endpoint creates / updates:\n"
         "  - product.template               (PRODUCTS sheet)\n"
         "  - product.template.attribute.line  (ATTRIBUTE_LINES sheet)\n"
         "  - product.template.attribute.value (ATTRIBUTE_VALUES sheet - price_extra overrides)\n"
         "  - mrp.bom + mrp.bom.line          (BOM_HEADERS + BOM_LINES)\n"
         "  - Hardware components             (HARDWARE_BOM)\n"
         "  - Accessory option products       (ACCESSORIES)\n\n"
         "All reference sheets (REF_*) are READ-ONLY dropdown sources. Don't\n"
         "edit them - the importer reads them to validate row data."),
        ("", None),
        ("WORKFLOW", SUB_FONT),
        (None,
         "1. Open this file in Excel (or LibreOffice Calc / Google Sheets).\n"
         "2. Fill the PRODUCTS sheet - at least one row.\n"
         "3. For each product that uses the configurator, add its attributes\n"
         "   to ATTRIBUTE_LINES (one row per attribute on the product).\n"
         "4. For per-template attribute pricing overrides, add ATTRIBUTE_VALUES rows.\n"
         "5. For manufactured products, add a BOM_HEADERS row + N BOM_LINES.\n"
         "6. Hardware (hinges, slides, legs) go in HARDWARE_BOM.\n"
         "7. Accessory option products go in ACCESSORIES.\n"
         "8. Save the file, upload via:  Backend -> Southbrook -> Import\n"
         "   Products -> drop this file.  Preview shows every row with VALID /\n"
         "   ERROR status before anything is written.\n"
         "9. Click 'Commit valid rows' on the preview modal to write to Odoo.\n"
         "   Errored rows are skipped; download the error report for fixup."),
        ("", None),
        ("KEY RULES", SUB_FONT),
        (None,
         "- default_code (SKU) is the primary key. Importer upserts by SKU.\n"
         "- External ID (xml_id) is optional but recommended for stable refs.\n"
         "- Linked sheets join on SKU. Example: BOM_LINES.parent_sku matches\n"
         "  a PRODUCTS.default_code row OR a BOM_HEADERS.parent_sku row.\n"
         "- Blank cells use the importer default (e.g. blank uom_id -> 'Units').\n"
         "- Booleans accept TRUE / FALSE / yes / no / 1 / 0 (case-insensitive).\n"
         "- Prices in CAD (Southbrook stack default).\n"
         "- Configurator rules (e.g. White Melamine forbids wood stains) live\n"
         "  in the rule engine, NOT this spreadsheet. The importer validates\n"
         "  attribute values against the live REF_ATTRIBUTES vocabulary."),
        ("", None),
        ("WHERE THE LIVE VOCABULARY COMES FROM", SUB_FONT),
        (None,
         "The REF_* sheets in this template were generated from a snapshot of\n"
         "the live Odoo DB on the QNAP container. Specifically:\n"
         "  - REF_CATEGORIES  ->  product.category records  (backend /odoo/product-categories)\n"
         "  - REF_ATTRIBUTES  ->  product.attribute + product.attribute.value\n"
         "                       records (backend /odoo/attributes)\n"
         "  - REF_UOM         ->  uom.uom records\n"
         "To regenerate this template against the latest vocab, re-run:\n"
         "  scripts/gen_import_template.py --refresh"),
        ("", None),
        ("REVIEWED FIELDS", SUB_FONT),
        (None, "See REF_FIELDS for the column -> Odoo field crosswalk."),
        ("", None),
        ("BACKEND ACTION SHORTCUTS", SUB_FONT),
        (None,
         "  - Products (product.template)   ->  /odoo/action-643\n"
         "  - Product Variants               ->  /odoo/action-644\n"
         "  - Attribute Values               ->  /odoo/action-1026\n"
         "  - Product Categories             ->  /odoo/product-categories\n"
         "  - Attributes (vocabulary)        ->  /odoo/attributes"),
    ]
    r = 1
    for label, body in intro:
        if isinstance(label, str) and label and body is None:
            c = ws.cell(row=r, column=1, value=label)
            c.font = TITLE_FONT if label == intro[0][0] else SUB_FONT
            ws.cell(row=r, column=2)
        elif label is None and body:
            c = ws.cell(row=r, column=2, value=body)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = NOTE_FONT
            ws.row_dimensions[r].height = max(15, body.count("\n") * 14 + 16)
        r += 1
    ws.cell(row=1, column=2, value="(see right column for the full guide)")

    # -------- PRODUCTS --------
    ws = wb.create_sheet("PRODUCTS")
    set_header(ws, PRODUCT_COLS)
    for i, row in enumerate(cabinets[:12], 2):
        xid, _id, code, nm, lp, scat, sdim, sicon, sdesc = row
        code_or_dash = code or f"SB-{xid.upper().replace('_','-')}"
        ws.cell(row=i, column=1,  value=f"southbrook_estimating.{xid}")
        ws.cell(row=i, column=2,  value=code_or_dash)
        ws.cell(row=i, column=3,  value=nm)
        ws.cell(row=i, column=4,  value="consu")
        ws.cell(row=i, column=5,  value="TRUE")
        ws.cell(row=i, column=6,  value="Goods")
        ws.cell(row=i, column=7,  value="Units")
        ws.cell(row=i, column=8,  value=float(lp))
        ws.cell(row=i, column=10, value="TRUE")
        ws.cell(row=i, column=11, value="TRUE")
        ws.cell(row=i, column=12, value="TRUE")
        ws.cell(row=i, column=13, value="FALSE")
        ws.cell(row=i, column=14, value="TRUE")
        ws.cell(row=i, column=18, value=scat)
        ws.cell(row=i, column=19, value=sdesc)
        ws.cell(row=i, column=20, value=sdim)
        ws.cell(row=i, column=21, value=sicon)
        ws.cell(row=i, column=22, value="TRUE")
        for c_idx in range(1, len(PRODUCT_COLS) + 1):
            ws.cell(row=i, column=c_idx).fill = INPUT_FILL

    # -------- ATTRIBUTE_LINES --------
    ws = wb.create_sheet("ATTRIBUTE_LINES")
    set_header(ws, ATTR_LINE_COLS)
    for i, (sku, attr, vals) in enumerate(EXAMPLE_ATTR_LINES, 2):
        ws.cell(row=i, column=1, value=sku)
        ws.cell(row=i, column=2, value=attr)
        ws.cell(row=i, column=3, value=vals)
        for c_idx in range(1, len(ATTR_LINE_COLS) + 1):
            ws.cell(row=i, column=c_idx).fill = INPUT_FILL

    # -------- ATTRIBUTE_VALUES --------
    ws = wb.create_sheet("ATTRIBUTE_VALUES")
    set_header(ws, ATTR_VAL_COLS)
    for i, vals in enumerate(EXAMPLE_ATTR_VALUES, 2):
        for c_idx, v in enumerate(vals, 1):
            ws.cell(row=i, column=c_idx, value=v).fill = INPUT_FILL

    # -------- BOM_HEADERS --------
    ws = wb.create_sheet("BOM_HEADERS")
    set_header(ws, BOM_HEAD_COLS)
    ws.cell(row=2, column=1, value="SB-BASE-1DR")
    ws.cell(row=2, column=2, value="SB-BASE-1DR BOM")
    ws.cell(row=2, column=3, value="normal")
    ws.cell(row=2, column=4, value=1)
    ws.cell(row=2, column=5, value="Units")
    for c_idx in range(1, len(BOM_HEAD_COLS) + 1):
        ws.cell(row=2, column=c_idx).fill = INPUT_FILL

    # -------- BOM_LINES --------
    ws = wb.create_sheet("BOM_LINES")
    set_header(ws, BOM_LINE_COLS)
    for i, vals in enumerate(EXAMPLE_BOM_LINES, 2):
        for c_idx, v in enumerate(vals, 1):
            ws.cell(row=i, column=c_idx, value=v).fill = INPUT_FILL

    # -------- HARDWARE_BOM --------
    ws = wb.create_sheet("HARDWARE_BOM")
    set_header(ws, HW_COLS)
    for i, vals in enumerate(EXAMPLE_HW, 2):
        for c_idx, v in enumerate(vals, 1):
            ws.cell(row=i, column=c_idx, value=v).fill = INPUT_FILL

    # -------- ACCESSORIES --------
    ws = wb.create_sheet("ACCESSORIES")
    set_header(ws, ACC_COLS)
    for i, vals in enumerate(EXAMPLE_ACC, 2):
        for c_idx, v in enumerate(vals, 1):
            ws.cell(row=i, column=c_idx, value=v).fill = INPUT_FILL

    # -------- REF_CATEGORIES --------
    ws = wb.create_sheet("REF_CATEGORIES")
    set_header(ws, [("id", "Odoo category id"), ("name", "Use in PRODUCTS.internal_category")])
    for i, (cid, cname) in enumerate(categs, 2):
        ws.cell(row=i, column=1, value=int(cid)).fill = REF_FILL
        ws.cell(row=i, column=2, value=cname).fill = REF_FILL

    # -------- REF_ATTRIBUTES --------
    ws = wb.create_sheet("REF_ATTRIBUTES")
    set_header(ws, [
        ("attribute_name", "Attribute name (use in ATTRIBUTE_LINES.attribute_name)"),
        ("display_type",   "OCA display_type: select / radio / color / multi"),
        ("create_variant", "When variants are created: dynamic / always / no_variant"),
        ("valid_values",   "Pipe-separated valid values"),
    ])
    for i, row in enumerate(attrs, 2):
        name, dtype, cvar, vals = row
        ws.cell(row=i, column=1, value=name).fill = REF_FILL
        ws.cell(row=i, column=2, value=dtype).fill = REF_FILL
        ws.cell(row=i, column=3, value=cvar).fill = REF_FILL
        ws.cell(row=i, column=4, value=vals or "").fill = REF_FILL

    # -------- REF_UOM --------
    ws = wb.create_sheet("REF_UOM")
    set_header(ws, [("id", "Odoo uom id"), ("name", "Use in PRODUCTS.uom_id / BOM_LINES.product_uom")])
    for i, (uid, uname) in enumerate(uoms, 2):
        ws.cell(row=i, column=1, value=int(uid)).fill = REF_FILL
        ws.cell(row=i, column=2, value=uname).fill = REF_FILL

    # -------- REF_FIELDS --------
    ws = wb.create_sheet("REF_FIELDS")
    set_header(ws, [
        ("sheet",        "This template's sheet name"),
        ("template_col", "Column header in the template"),
        ("odoo_model",   "Odoo model the importer writes to"),
        ("odoo_field",   "Odoo field name"),
        ("writable_via", "How the importer resolves / writes: direct / name lookup / xml_id ref / computed"),
        ("phase_added",  "Which Phase added this field - for change history"),
    ])
    for i, vals in enumerate(FIELD_MAP, 2):
        for c_idx, v in enumerate(vals, 1):
            ws.cell(row=i, column=c_idx, value=v).fill = REF_FILL

    # -------- VERSION_STAMP --------
    ws = wb.create_sheet("VERSION_STAMP")
    rows = [
        ("Template version",          VERSION),
        ("Generated at",               generated_at),
        ("Source",                     "scripts/gen_import_template.py against live southbrook-odoo container at QNAP 192.168.68.108"),
        ("Importer endpoint",          "POST /southbrook/api/import/preview  (preview)  +  POST /southbrook/api/import/commit  (commit, requires confirm:true)"),
        ("Expected importer version",  "Phase 4 endpoint - see southbrook_configurator_ux Phase 4 in the module manifest"),
    ]
    for r, (k, v) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100

    # -------- Data validations --------
    add_validations(wb, len(categs), len(uoms), len(attrs))

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out, len(wb.sheetnames)


def add_validations(wb, n_categs, n_uoms, n_attrs):
    """Attach dropdown validations to data-entry columns."""

    cat_last = n_categs + 1
    uom_last = n_uoms + 1
    attr_last = n_attrs + 1

    products = wb["PRODUCTS"]
    products.add_data_validation(_dv_list_ref(f"=REF_CATEGORIES!$B$2:$B${cat_last}",
                                              "Pick a category from REF_CATEGORIES (or add it in Odoo first).",
                                              "Unknown category"))
    products.data_validations.dataValidation[-1].add("F2:F500")

    products.add_data_validation(_dv_list_ref(f"=REF_UOM!$B$2:$B${uom_last}"))
    products.data_validations.dataValidation[-1].add("G2:G500")

    products.add_data_validation(_dv_list_inline("Wall,Base,Drawer,Tall,Vanity,Extras"))
    products.data_validations.dataValidation[-1].add("R2:R500")

    products.add_data_validation(_dv_list_inline(
        "wall1,wall2,base1,base2,drawer,sink,pantry,oven,corner,vanity,extra,worktop"))
    products.data_validations.dataValidation[-1].add("U2:U500")

    products.add_data_validation(_dv_list_inline("TRUE,FALSE"))
    for col in ("E", "J", "K", "L", "M", "N", "V"):
        products.data_validations.dataValidation[-1].add(f"{col}2:{col}500")

    products.add_data_validation(_dv_list_inline("consu,service,combo"))
    products.data_validations.dataValidation[-1].add("D2:D500")

    attr_lines = wb["ATTRIBUTE_LINES"]
    attr_lines.add_data_validation(_dv_list_ref(f"=REF_ATTRIBUTES!$A$2:$A${attr_last}"))
    attr_lines.data_validations.dataValidation[-1].add("B2:B500")
    attr_lines.add_data_validation(_dv_list_inline("TRUE,FALSE"))
    attr_lines.data_validations.dataValidation[-1].add("E2:E500")

    attr_vals = wb["ATTRIBUTE_VALUES"]
    attr_vals.add_data_validation(_dv_list_ref(f"=REF_ATTRIBUTES!$A$2:$A${attr_last}"))
    attr_vals.data_validations.dataValidation[-1].add("B2:B500")

    bom_head = wb["BOM_HEADERS"]
    bom_head.add_data_validation(_dv_list_inline("normal,phantom,subcontract"))
    bom_head.data_validations.dataValidation[-1].add("C2:C500")
    bom_head.add_data_validation(_dv_list_ref(f"=REF_UOM!$B$2:$B${uom_last}"))
    bom_head.data_validations.dataValidation[-1].add("E2:E500")

    bom_lines = wb["BOM_LINES"]
    bom_lines.add_data_validation(_dv_list_ref(f"=REF_UOM!$B$2:$B${uom_last}"))
    bom_lines.data_validations.dataValidation[-1].add("E2:E500")
    bom_lines.add_data_validation(_dv_list_inline("Length,Width,None"))
    bom_lines.data_validations.dataValidation[-1].add("J2:J500")
    bom_lines.add_data_validation(_dv_list_inline("TRUE,FALSE"))
    bom_lines.data_validations.dataValidation[-1].add("K2:K500")

    hw = wb["HARDWARE_BOM"]
    hw.add_data_validation(_dv_list_inline(
        "Hinge,Drawer Slide,Leg,Leg Clip,Toe-Kick Clip,Shelf Pin,Handle,Pin,Connector,Knob"))
    hw.data_validations.dataValidation[-1].add("B2:B500")
    hw.add_data_validation(_dv_list_ref(f"=REF_UOM!$B$2:$B${uom_last}"))
    hw.data_validations.dataValidation[-1].add("G2:G500")

    acc = wb["ACCESSORIES"]
    acc.add_data_validation(_dv_list_inline(
        "End Panel,Filler,Cornice,Pelmet,Plinth,Internal Pull-Out,Soft-Close,Drawer Organisers"))
    acc.data_validations.dataValidation[-1].add("C2:C500")
    acc.add_data_validation(_dv_list_inline("TRUE,FALSE"))
    acc.data_validations.dataValidation[-1].add("F2:F500")


def _dv_list_inline(values: str) -> DataValidation:
    return DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)


def _dv_list_ref(formula: str, error: str | None = None,
                 title: str | None = None) -> DataValidation:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    if error:
        dv.error = error
    if title:
        dv.errorTitle = title
    return dv


# =====================================================================
# CLI
# =====================================================================

def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                      formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT,
                        help=f"output xlsx path (default: {DEFAULT_OUT})")
    parser.add_argument("--refresh", action="store_true",
                        help="re-pull live vocab from the QNAP container before building")
    args = parser.parse_args()

    if args.refresh:
        refresh_vocab()

    attrs = load_tsv(ATTRS_TSV)
    categs = load_tsv(CATEGS_TSV)
    uoms = load_tsv(UOM_TSV)
    cabinets = load_tsv(CABINETS_TSV)

    out, n_sheets = build(args.out, attrs, categs, uoms, cabinets)
    size = os.path.getsize(out)
    sys.stderr.write(f"Wrote {out}  ({size:,} bytes, {n_sheets} sheets)\n")


if __name__ == "__main__":
    main()
